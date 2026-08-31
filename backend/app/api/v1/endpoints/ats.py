from typing import Any, List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
import pandas as pd
import io
import re

from app.core.database import get_db
from app.models.student import Student
from app.models.placement_drive import PlacementDrive

router = APIRouter()

class JobDescriptionInput(BaseModel):
    company_name: str = "Target Company"
    role_title: str = "Software Development Engineer"
    required_skills: List[str] = []
    min_cgpa: float = 6.0
    eligible_departments: List[str] = []
    description_text: Optional[str] = None

class CandidateInput(BaseModel):
    id: Optional[int] = None
    roll_number: str
    name: str
    department: str
    cgpa: float = 0.0
    skills: List[str] = []
    resume_link: Optional[str] = None
    email: Optional[str] = None
    mobile_number: Optional[str] = None
    raw_text: Optional[str] = None

class CompareRequest(BaseModel):
    job_description: JobDescriptionInput
    candidates: Optional[List[CandidateInput]] = None
    use_db_students: bool = False

COMMON_TECH_SKILLS = [
    "python", "java", "c++", "c#", "c", "javascript", "typescript", "react", "next.js", "angular", "vue",
    "node.js", "express", "fastapi", "django", "flask", "spring boot", "sql", "postgresql", "mysql",
    "mongodb", "redis", "docker", "kubernetes", "aws", "azure", "gcp", "git", "github", "linux",
    "html", "css", "tailwind", "rest api", "graphql", "machine learning", "deep learning", "nlp",
    "data structures", "algorithms", "dsa", "oop", "system design", "microservices", "ci/cd",
    "tableau", "power bi", "pandas", "numpy", "pytorch", "tensorflow", "agile", "devops"
]

def extract_skills_from_text(text: str) -> List[str]:
    if not text:
        return []
    text_lower = text.lower()
    found_skills = []
    for skill in COMMON_TECH_SKILLS:
        pattern = r'(?:\b|_)' + re.escape(skill) + r'(?:\b|_)'
        if re.search(pattern, text_lower):
            found_skills.append(skill.title() if len(skill) > 3 else skill.upper())
    return found_skills

def get_department_skills(dept: str) -> List[str]:
    d = (dept or "").strip().lower()
    if "cyber" in d or "security" in d:
        return ["Network Security", "Cryptography", "Linux", "Python", "Ethical Hacking", "SQL", "Git", "C++", "Wireshark"]
    elif "info" in d or "it" in d or "computer" in d or "cse" in d or "software" in d:
        return ["Python", "Java", "SQL", "Data Structures", "Algorithms", "React", "Node.js", "Git", "C++", "System Design"]
    elif "ai" in d or "data" in d:
        return ["Python", "Machine Learning", "Data Analysis", "SQL", "Pandas", "PyTorch", "TensorFlow", "Deep Learning", "Git"]
    elif "electronic" in d or "ece" in d or "communi" in d:
        return ["Embedded Systems", "IoT", "C", "C++", "MATLAB", "Linux", "Microcontrollers", "Python"]
    elif "electri" in d or "eee" in d:
        return ["Power Systems", "Control Systems", "MATLAB", "C", "Circuit Design", "Python"]
    elif "mech" in d:
        return ["AutoCAD", "SolidWorks", "Finite Element Analysis", "Thermodynamics", "Python", "ANSYS"]
    elif "business" in d or "bba" in d or "management" in d or "commerce" in d:
        return ["Business Analysis", "Financial Analysis", "Project Management", "MS Excel", "SQL", "Data Analysis", "Communication"]
    else:
        return ["Problem Solving", "Python", "SQL", "Data Structures", "Git"]

def is_department_match(cand_dept: str, eligible_depts: List[str]) -> bool:
    if not eligible_depts or "ALL" in [d.upper() for d in eligible_depts]:
        return True
    c_lower = cand_dept.strip().lower()
    for req in eligible_depts:
        r_lower = req.strip().lower()
        if r_lower in c_lower or c_lower in r_lower:
            return True
        # Abbreviations mapping
        if r_lower == "cse" and ("computer" in c_lower or "cse" in c_lower or "cyber" in c_lower or "software" in c_lower):
            return True
        if r_lower == "it" and ("info" in c_lower or "it" in c_lower or "cyber" in c_lower):
            return True
        if r_lower == "aids" and ("ai" in c_lower or "data" in c_lower):
            return True
        if r_lower == "ece" and ("electronic" in c_lower or "ece" in c_lower):
            return True
        if r_lower == "eee" and ("electri" in c_lower or "eee" in c_lower):
            return True
        if r_lower == "bba" and ("business" in c_lower or "management" in c_lower or "bba" in c_lower):
            return True
    return False

def calculate_match(candidate_data: dict, jd: JobDescriptionInput) -> dict:
    jd_skills = [s.strip().lower() for s in jd.required_skills if s.strip()]
    cand_skills = [s.strip().lower() for s in candidate_data.get("skills", []) if s.strip()]
    
    # Also check cand raw_text or department for implicit skills
    if candidate_data.get("raw_text"):
        extra_skills = [s.lower() for s in extract_skills_from_text(candidate_data["raw_text"])]
        cand_skills = list(set(cand_skills + extra_skills))
        
    # Department default skills
    if not cand_skills:
        dept_skills = [s.lower() for s in get_department_skills(candidate_data.get("department", ""))]
        cand_skills = list(set(cand_skills + dept_skills))

    matched_skills = []
    missing_skills = []
    
    for req in jd_skills:
        if any(req in cs or cs in req for cs in cand_skills):
            matched_skills.append(req.title())
        else:
            missing_skills.append(req.title())
            
    # Skill Match Score (50% weight)
    if len(jd_skills) > 0:
        skill_score = (len(matched_skills) / len(jd_skills)) * 100.0
    else:
        skill_score = 80.0

    # CGPA Score (25% weight)
    raw_cgpa = float(candidate_data.get("cgpa") or 0.0)
    ug_pct = float(candidate_data.get("ug_percentage") or 0.0)
    if raw_cgpa <= 1.0 and ug_pct > 10.0:
        cgpa = round(ug_pct / 10.0, 2)
    else:
        cgpa = raw_cgpa if raw_cgpa > 0 else 7.5

    min_cgpa = float(jd.min_cgpa or 6.0)
    if cgpa >= min_cgpa:
        cgpa_score = 100.0
        cgpa_eligible = True
    elif cgpa >= (min_cgpa - 0.5):
        cgpa_score = 75.0
        cgpa_eligible = False
    else:
        cgpa_score = max(0.0, (cgpa / max(min_cgpa, 1.0)) * 60.0)
        cgpa_eligible = False

    # Department Match (15% weight)
    cand_dept = str(candidate_data.get("department", "")).strip()
    dept_eligible = is_department_match(cand_dept, jd.eligible_departments)
    dept_score = 100.0 if dept_eligible else 45.0

    # Profile Completeness (10% weight)
    profile_score = 100.0 if candidate_data.get("resume_link") else 70.0

    # Total Overall Weighted Score
    overall_match = (
        (skill_score * 0.50) +
        (cgpa_score * 0.25) +
        (dept_score * 0.15) +
        (profile_score * 0.10)
    )
    overall_match = round(min(100.0, max(0.0, overall_match)), 1)

    # Recommendation category
    if overall_match >= 80 and cgpa_eligible and dept_eligible:
        recommendation = "Strong Match"
        status_color = "emerald"
    elif overall_match >= 65:
        recommendation = "Good Match"
        status_color = "blue"
    elif overall_match >= 50:
        recommendation = "Potential Match"
        status_color = "amber"
    else:
        recommendation = "Low Match"
        status_color = "rose"

    return {
        "candidate": candidate_data,
        "match_score": overall_match,
        "skill_score": round(skill_score, 1),
        "cgpa_score": round(cgpa_score, 1),
        "dept_score": round(dept_score, 1),
        "matched_skills": matched_skills,
        "missing_skills": missing_skills,
        "cgpa_eligible": cgpa_eligible,
        "dept_eligible": dept_eligible,
        "recommendation": recommendation,
        "status_color": status_color,
        "analysis": f"Matched {len(matched_skills)}/{len(jd_skills)} required skills. CGPA: {cgpa} vs {min_cgpa} cutoff."
    }

@router.post("/compare")
def compare_ats(payload: CompareRequest, db: Session = Depends(get_db)):
    candidates_list = []
    
    # If use_db_students or no candidates passed, pull from existing DB
    if payload.use_db_students or not payload.candidates:
        db_students = db.query(Student).filter(Student.is_deleted == False).all()
        for s in db_students:
            # CGPA conversion if 0
            effective_cgpa = s.cgpa if s.cgpa > 1.0 else (round(s.ug_percentage / 10.0, 2) if s.ug_percentage and s.ug_percentage > 10 else 7.5)
            
            # Build skills from links/department
            skills = get_department_skills(s.department)
            if s.github_link:
                skills.extend(["Git", "GitHub"])
            if s.portfolio_link:
                skills.extend(["Web Development", "React"])
                
            candidates_list.append({
                "id": s.id,
                "roll_number": s.roll_number,
                "name": s.name,
                "department": s.department,
                "cgpa": effective_cgpa,
                "ug_percentage": s.ug_percentage,
                "skills": list(set(skills)),
                "resume_link": s.resume_link,
                "github_link": s.github_link,
                "linkedin_link": s.linkedin_link,
                "portfolio_link": s.portfolio_link,
                "email": s.email,
                "mobile_number": s.mobile_number,
                "placement_status": s.placement_status
            })
    else:
        for c in payload.candidates:
            candidates_list.append(c.model_dump())

    results = []
    for cand in candidates_list:
        match_result = calculate_match(cand, payload.job_description)
        results.append(match_result)

    # Sort descending by match score
    results.sort(key=lambda x: x["match_score"], reverse=True)

    # Summary statistics
    total = len(results)
    strong_matches = sum(1 for r in results if r["match_score"] >= 80)
    good_matches = sum(1 for r in results if 65 <= r["match_score"] < 80)
    avg_score = round(sum(r["match_score"] for r in results) / total, 1) if total > 0 else 0

    return {
        "job_description": payload.job_description.model_dump(),
        "total_analyzed": total,
        "strong_matches": strong_matches,
        "good_matches": good_matches,
        "average_score": avg_score,
        "results": results
    }

@router.post("/parse-excel")
async def parse_resumes_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Please upload a valid Excel or CSV file")
        
    try:
        contents = await file.read()
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents))
            
        df = df.dropna(how='all')
        df = df.astype(object).where(pd.notnull(df), None)
        
        candidates = []
        for idx, row in df.iterrows():
            row_dict = {str(k).strip().lower().replace(" ", "_"): v for k, v in row.items() if v is not None}
            
            # Map columns
            roll_number = str(row_dict.get("roll_number") or row_dict.get("roll_no") or row_dict.get("reg_no") or f"STU{idx+1:03d}")
            name = str(row_dict.get("name") or row_dict.get("student_name") or row_dict.get("candidate_name") or f"Candidate {idx+1}")
            department = str(row_dict.get("department") or row_dict.get("dept") or row_dict.get("branch") or "CSE")
            
            try:
                cgpa = float(row_dict.get("cgpa") or row_dict.get("gpa") or 0.0)
            except:
                cgpa = 0.0
                
            skills_raw = row_dict.get("skills") or row_dict.get("technical_skills") or row_dict.get("key_skills") or ""
            if isinstance(skills_raw, str):
                skills = [s.strip() for s in re.split(r'[,;|\n]+', skills_raw) if s.strip()]
            elif isinstance(skills_raw, list):
                skills = skills_raw
            else:
                skills = []
                
            # If no skills in column, auto-extract from projects/experience or raw columns
            combined_text = " ".join([str(v) for v in row.values if v is not None])
            if not skills:
                skills = extract_skills_from_text(combined_text)
                
            resume_link = str(row_dict.get("resume_link") or row_dict.get("resume") or row_dict.get("portfolio") or "")
            email = str(row_dict.get("email") or row_dict.get("email_id") or "")
            phone = str(row_dict.get("mobile_number") or row_dict.get("phone") or row_dict.get("contact") or "")
            
            candidates.append({
                "id": idx + 1,
                "roll_number": roll_number,
                "name": name,
                "department": department,
                "cgpa": cgpa,
                "skills": skills,
                "resume_link": resume_link if resume_link.startswith("http") else None,
                "email": email,
                "mobile_number": phone,
                "raw_text": combined_text[:300]
            })
            
        return {
            "file_name": file.filename,
            "total_candidates": len(candidates),
            "candidates": candidates
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to parse resume Excel file: {str(e)}")

@router.post("/parse-jd")
def parse_job_description(payload: dict):
    """
    Parse raw Job Description text to extract company, role, skills, CGPA cutoff, and eligible branches.
    """
    raw_text = payload.get("jd_text", "")
    if not raw_text:
        raise HTTPException(status_code=400, detail="Job description text is required")
        
    extracted_skills = extract_skills_from_text(raw_text)
    
    # Try to extract CGPA
    cgpa_match = re.search(r'(?:cgpa|gpa|percentage|pointer)\s*(?:of|above|>=|>|minimum|min)?\s*([0-9]+(?:\.[0-9]+)?)', raw_text, re.IGNORECASE)
    min_cgpa = 6.5
    if cgpa_match:
        try:
            val = float(cgpa_match.group(1))
            if val <= 10.0:
                min_cgpa = val
            elif val <= 100.0:
                min_cgpa = round(val / 10.0, 1)
        except:
            pass

    # Try to extract departments
    dept_map = {
        "CSE": ["cse", "computer science", "computer engineering"],
        "IT": ["it", "information technology"],
        "AIDS": ["aids", "ai", "artificial intelligence", "data science"],
        "ECE": ["ece", "electronics", "communication"],
        "EEE": ["eee", "electrical"],
        "MECH": ["mech", "mechanical"]
    }
    found_depts = []
    text_lower = raw_text.lower()
    for dept, keywords in dept_map.items():
        if any(k in text_lower for k in keywords):
            found_depts.append(dept)
    if not found_depts:
        found_depts = ["CSE", "IT", "AIDS", "ECE"]

    return {
        "company_name": payload.get("company_name", "Target Company"),
        "role_title": payload.get("role_title", "Software Engineer"),
        "required_skills": extracted_skills,
        "min_cgpa": min_cgpa,
        "eligible_departments": found_depts,
        "description_text": raw_text
    }

@router.get("/template")
def download_ats_template():
    columns = [
        "Roll Number", "Name", "Department", "CGPA", "Skills", 
        "Resume Link", "Email", "Mobile Number", "Projects Summary"
    ]
    sample_data = [
        ["STU101", "Arjun Sharma", "CSE", 8.8, "Python, React, Node.js, SQL, AWS, Docker", "https://example.com/resume1.pdf", "arjun@example.com", "9876543210", "Fullstack Ecommerce & AI Chatbot"],
        ["STU102", "Priya Verma", "IT", 9.2, "Java, Spring Boot, Microservices, PostgreSQL, Kubernetes", "https://example.com/resume2.pdf", "priya@example.com", "9876543211", "Banking microservice backend"],
        ["STU103", "Karan Reddy", "ECE", 8.4, "C++, Python, Embedded Systems, IoT, Linux", "https://example.com/resume3.pdf", "karan@example.com", "9876543212", "Smart Home IoT Controller"],
        ["STU104", "Sneha Patel", "AIDS", 9.0, "Python, Machine Learning, TensorFlow, PyTorch, SQL, Pandas", "https://example.com/resume4.pdf", "sneha@example.com", "9876543213", "Healthcare Diagnostic ML Model"],
        ["STU105", "Vikram Das", "CSE", 7.9, "JavaScript, HTML, CSS, React, MongoDB, Express", "https://example.com/resume5.pdf", "vikram@example.com", "9876543214", "Social Media Web Platform"],
        ["STU106", "Ananya Iyer", "CSE", 9.4, "Python, FastAPI, React, TypeScript, Docker, GCP", "https://example.com/resume6.pdf", "ananya@example.com", "9876543215", "Cloud-native Microservices Analytics"],
        ["STU107", "Rohan Gupta", "IT", 8.1, "Java, C++, DSA, SQL, Git, Linux", "https://example.com/resume7.pdf", "rohan@example.com", "9876543216", "High performance distributed cache"],
        ["STU108", "Meera Joshi", "AIDS", 8.9, "Python, NLP, PyTorch, Deep Learning, SQL", "https://example.com/resume8.pdf", "meera@example.com", "9876543217", "GenAI Document Summarizer"]
    ]
    
    from openpyxl.utils import get_column_letter
    df = pd.DataFrame(sample_data, columns=columns)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='ATS_Candidates_Template', index=False)
        worksheet = writer.sheets['ATS_Candidates_Template']
        for idx, col in enumerate(df.columns, 1):
            worksheet.column_dimensions[get_column_letter(idx)].width = 24
            
    output.seek(0)
    headers = {'Content-Disposition': 'attachment; filename="ats_student_resumes_template.xlsx"'}
    return StreamingResponse(
        output,
        headers=headers,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@router.post("/export-results")
def export_comparison_results(payload: dict):
    results = payload.get("results", [])
    jd = payload.get("job_description", {})
    
    export_rows = []
    for r in results:
        cand = r.get("candidate", {})
        export_rows.append({
            "Rank": len(export_rows) + 1,
            "Roll Number": cand.get("roll_number", ""),
            "Name": cand.get("name", ""),
            "Department": cand.get("department", ""),
            "CGPA": cand.get("cgpa", 0.0),
            "Match Score (%)": r.get("match_score", 0),
            "Recommendation": r.get("recommendation", ""),
            "CGPA Eligible": "Yes" if r.get("cgpa_eligible") else "No",
            "Dept Eligible": "Yes" if r.get("dept_eligible") else "No",
            "Matched Skills": ", ".join(r.get("matched_skills", [])),
            "Missing Skills": ", ".join(r.get("missing_skills", [])),
            "All Skills": ", ".join(cand.get("skills", [])),
            "Email": cand.get("email", ""),
            "Phone": cand.get("mobile_number", ""),
            "Resume Link": cand.get("resume_link", "")
        })

    from openpyxl.utils import get_column_letter
    df = pd.DataFrame(export_rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='ATS_Match_Results', index=False)
        worksheet = writer.sheets['ATS_Match_Results']
        for idx, col in enumerate(df.columns, 1):
            worksheet.column_dimensions[get_column_letter(idx)].width = 20

    output.seek(0)
    filename = f"ATS_Results_{jd.get('company_name', 'Company').replace(' ', '_')}.xlsx"
    headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        headers=headers,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

